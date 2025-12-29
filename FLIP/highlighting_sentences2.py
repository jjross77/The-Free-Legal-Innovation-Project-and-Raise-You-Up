# -*- coding: utf-8 -*-
"""
Created on Sun Mar 12 19:22:50 2023

@author: yyyyyyyyyyyyyyyyyyyy
"""

import pickle
import re
from openpyxl import load_workbook

green = "#00FF00" 	
blue = "#00FFFF"
orange = "#FFA500"	
yellow = "#FFFF00"	
pink =  "#FF00FF"	
deep_sky_blue = "#00BFFF"	
sandy_brown = "#F4A460"	
sea_green = "#3CB371"		
mint_cream_blue = "#F5FFFA"	
red = "#FF0000"
purple = "#7F00FF"
dark_blue = "#00008B"
bright_red = "#EE4B2B"
bright_purple = "#BF40BF"
dark_green = "#006400"
counter=0
list_of_labels_2 = []
list_of_values = []
sentence_list = []

dictionary_of_colors= {0: bright_red , 1: orange, 2: yellow , 3 : green , 4 : blue, 5 : bright_purple ,\
                       6 : pink, 7: mint_cream_blue, 8 : deep_sky_blue , 9 :sea_green }





path = r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\Kimlichcova\Sparrow_with_dic_for_showcase.xlsx"

wb = load_workbook(filename=path)
sheet = wb.active
storing = 0
for  value in sheet['A']:
    storing += 1
    if value.value == None:
        break 
storing= storing+1 
sub_cell_list = []
obj_cell_list = []
verb_cell_list = []
pos_cell_list = []
for iiiii in range(2, storing):
    
        pos_cell = sheet.cell(row=iiiii, column = 2).value
        sub_cell = sheet.cell(row=iiiii, column = 3).value
        obj_cell = sheet.cell(row=iiiii, column = 4).value
        verb_cell = sheet.cell(row=iiiii, column = 5).value
        sub_cell_list.append(sub_cell)
        obj_cell_list.append(obj_cell)
        verb_cell_list.append(verb_cell)
        pos_cell_list.append(pos_cell)



                	
with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\numpy_dict_of_embeddings.pickle","rb") as ff:
    original_case_dict=pickle.load(ff)
    
pattern_for_labels= re.compile(r'(\d+): (\d+): (\d+): (.*)') # group 2 is label 2 group 1 is label 1 group 4 is sentence



for key, value in original_case_dict.items():
 
    try:
        keyy=str(key)
        label_regex_obj=pattern_for_labels.search(keyy)
        sentence_group =label_regex_obj.group(4)
        #sentence_group= re.sub(' +', ' ', sentence_group)

        
        
        label_2=label_regex_obj.group(2)
        label_2=int(label_2)
        if label_2>10:
            if label_2==22:
                label_2=2
        
        
        list_of_labels_2.append(label_2)
        list_of_values.append(value)
        sentence_list.append(sentence_group)
        
        

    except Exception as e:
        print(e)
pos_cell_list_val = pos_cell_list[8500:]
sentence_list_val =    sentence_list[8500:]   
list_of_labels_2_val = list_of_labels_2[8500:]
#list_of_labels_2_val=np.array(list_of_labels_2_val)
list_of_values_val =  list_of_values[8500:]
sub_cell_list_val =  sub_cell_list[8500:]
obj_cell_list_val = obj_cell_list[8500:]
verb_cell_list_val =verb_cell_list[8500:]

Facts = 0
Analysis = 1
Holdings =2 
Issues = 3
Rules = 3
Conclusions = 4
Analysis_made_by_anyone_not_the_judge =5
Procedural_sentences = 6
NA = 7
Category_for_pieces_of_sentence = 8
Quotes = 9

with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\prediction_list.pickle", "rb") as f2:
    prediction_list=pickle.load(f2)
highlighted_sentence_list = []
with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\Kimlichcova\highlighted_case_and_highlighted_sentences.txt", "a", encoding = "utf-8") as f3:
    
    title_section = "<h5>" + "TAX THE RICH | TRANS RIGHTS ARE HUMAN RIGHTS | FIGHT FOR GREEN ENERGY | BLACK LIVES MATTER" + "</h5>"
    f3.write(title_section)
    Index = "<h2>" + "Index:" + "</h2>"
    f3.write(Index)
    index_font_meaning="<h3> 1: <span style='color: #330099;'>" + "Object Part" + "</span> 2: <span style='color: #339999;'>" + "Verb Part" + "</span> 3: <span style='color: #990000;'>" + "Subject Part" + "</span> </h3>"
    f3.write(index_font_meaning)
    index_sentence_fact=f"<h3 style=\"background-color: {dictionary_of_colors[Facts]}\">" + "Facts" + "</h3>"
    f3.write(index_sentence_fact)
    Analysiss =  f"<h3 style=\"background-color: {dictionary_of_colors[Analysis]}\">" + "Analysis" + "</h3>"
    f3.write(Analysiss)

    Holding =  f"<h3 style=\"background-color: {dictionary_of_colors[Holdings]}\">" + "Holdings" + "</h3>"
    f3.write(Holding)

    Issue =  f"<h3 style=\"background-color: {dictionary_of_colors[Issues]}\">" + "Issues" + "</h3>"
    f3.write(Issue)

    Rule =  f"<h3 style=\"background-color: {dictionary_of_colors[Rules]}\">" + "Rules" + "</h3>"
    f3.write(Rule)

    Conclusion =  f"<h3 style=\"background-color: {dictionary_of_colors[Conclusions]}\">" + "Conclusions" + "</h3>"
    f3.write(Conclusion)

    Analysis_by_others =  f"<h3 style=\"background-color: {dictionary_of_colors[Analysis_made_by_anyone_not_the_judge]}\">" + "Analysis Referenced" + "</h3>"
    f3.write(Analysis_by_others)

    Procedural =  f"<h3 style=\"background-color: {dictionary_of_colors[Procedural_sentences]}\">" + "Procedural Sentences" + "</h3>"
    f3.write(Procedural)

    NA_cat =  f"<h3 style=\"background-color: {dictionary_of_colors[NA]}\">" + "NA" + "</h3>"
    f3.write(NA_cat)

    Pieces_of_sentences =  f"<h3 style=\"background-color: {dictionary_of_colors[Category_for_pieces_of_sentence]}\">" + "Sentence Pieces" + "</h3>"
    f3.write(Pieces_of_sentences)

    Quote =  f"<h3 style=\"background-color: {dictionary_of_colors[Quotes]}\">" + "Quotes"  + "</h3>"
    f3.write(Quote)
    spaces =  "<h1>" + " <br> Case Sentences: "  + "</h1>"
    f3.write(spaces)

  
    
    
 
    for pos, prediction, subb, objj, verbb in zip(pos_cell_list_val, prediction_list, sub_cell_list_val, obj_cell_list_val, verb_cell_list_val): 
        temp_append_list_sub = []
        temp_append_list_obj = []
        temp_append_list_ver = []
        word_list = []
        pos = eval(pos)
        subb = eval(subb)
        objj=eval(objj)
        verbb= eval(verbb)
        for subb2 in subb:
            for subb3 in subb2:
                temp_append_list_sub.append(subb3[2])
                
        for obj2 in objj:
            for obj3 in obj2:
                temp_append_list_obj.append(obj3[2])
        for ver2 in verbb:
            for ver3 in ver2:
                temp_append_list_ver.append(ver3[2])
        
        for pos1 in pos:
            counterr = 0
            
            
            if pos1[2] in temp_append_list_obj:
                word ="<span style='color: #330099;'>" + pos1[0] + "</span>"
                counterr = 1

            if pos1[2] in temp_append_list_ver:
                word ="<span style='color: #339999;'>" + pos1[0] + "</span>"
                counterr = 1

            if pos1[2] in temp_append_list_sub:
                word ="<span style='color: #990000;'>" + pos1[0] + "</span>"
                counterr = 1

            if counterr == 0:
                word = pos1[0]
            
            word_list.append(word)
        word_sentence=" ".join(word_list)
        print(word_sentence)
        highlighted_sentence=f"<p style=\"background-color: {dictionary_of_colors[prediction]}\">" + word_sentence + "</p>"
        f3.write(highlighted_sentence)
            

        #if marked before then 
        # to use for object dark_green
        # to use for subject dark_blue
        # to use for verb red
        
        
        # do subject list last to write over object list if there is a copy
        
       