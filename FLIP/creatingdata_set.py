# -*- coding: utf-8 -*-
"""
Created on Sun Jan  8 23:27:19 2023

@author: rossg
"""
import numpy as np

from nltk.tokenize import word_tokenize


#Useful v Non-Useful

real_sentence = 1
non_sentence = 2



# FACTS

actual_facts = 1
lower_court_arg_parties = 2
lower_court_non_arg = 3
facts = actual_facts + lower_court_arg_parties + lower_court_non_arg

# Analysis

interpretation_analysis = 4
quotes_analysis = 5
tests_analysis = 6

analysis = interpretation_analysis + quotes_analysis + tests_analysis

#Issues

issue = 7
lower_court_issue = 8
issues = issue + lower_court_issue 


# Holdings
lower_court_holding = 9
current_court_holding = 10
holdings = lower_court_holding + current_court_holding

#token values

legislation = 1
jurisprudence = 2
articles = 3
coresponding_page_num = 4
references_leg = 5
references_juris = 6
references_article= 7

#Token Values for Interpretation Sentences
main_subject_being_interpreted = 1


# Type of Legal citiation

concurring = 1
varied_differentiated = 2 
overruled = 3


import openpyxl

from openpyxl import load_workbook
wb = load_workbook(filename=r'D:\Kimlichcova\sparrow_excelled.xlsx')
sheet = wb.active
Column_A = sheet['A'] 
Column_B = sheet['B']  # length that we start labeling again when we reopen the file. 
leng_of_columna= len(Column_A)
leng_of_columnB= len(Column_B)
wb = load_workbook(filename=r'D:\Kimlichcova\sparrow_excelled.xlsx')
sheet = wb.active
starting_value_minus2 = ""
storing = 0
for i, value in enumerate(sheet['B']):
    storing += 1
    if value.value == None:
        starting_value_minus2=value
        break  
starting_row_value= storing -1
print(starting_row_value)


helper_to_stopper = 0   
temp_list = []
sen_wait = 0
for  sentence in range(leng_of_columna):
    temp_list.clear()
    starting_row_value += 1
    c = sheet[f'A{starting_row_value}']
    split_sen=word_tokenize(c.value)
    print(c.value)
    len_sentence=len(split_sen)
    zero_list= [0 for i in range(len_sentence)]
    zero_list2= [0 for i in range(len_sentence)]

    for i, word in enumerate(split_sen):
        temp_item =f"{word}_({i})"
        temp_list.append(temp_item)
    
    
    print("\n ASSIGN A VALUE: WHOLE(1) or FRAGMENT(2) ")
    label1= int(input() )
    sheet[f'B{starting_row_value}'] = label1
    sheet[f'E{starting_row_value}']= str(split_sen)
    
    
    
    print("""ACTUAL_FACT_BY COURT = 1, lower_court_arg_parties = 1, lower_court_judge_non_arg = 1,
          INTERPRETATION_analysis = 4   Tests_analysis = 2,  Issue = 2,
          lower_court_issue = 2  lower_court_holding = 3 current_court_holding = 3
          Quotes_analysis = 5, N/A = 11""")
    print("\n Please assign Actual Fact(1), Interpretation (4) a value for whether is one of these values?")

    label2= int(input())
    sheet[f'C{starting_row_value}'] = label2
    
    
    
    
        
        
     
    print( "Is there Legislation or a Case in the sentence you need to assign here Y for yes or any other non-numeric value to answer no")
    has_useful_value = input()
    if has_useful_value == "y":
        
        print(zero_list)
        stopper= 0
        while stopper != "y": 
            
            print(temp_list)
                
                
            
            print(" Type y to quit.")
            print("CLICK ENTER OR SOME NON NUMERIC VALUE")
            
            stopper = input()
            if stopper =="y":
                break
            print("Select a word in the sentence to assign a value to.")
            select_word_in_sentence = int(input())
            print(split_sen[select_word_in_sentence])
            print(""" NA = 8, legislation = 1, jurisprudence = 2 , articles = 3, none_useful = 4
            references_leg = 5, references_juris = 6, references_subjects= 7""")
            print("Now assign the type of word that this is such as Jurisprudence(2) or Legislation(1), none_useful=4, references_leg = 5, references_juris = 6.")
            starting_point=zero_list[select_word_in_sentence]= int(input())
            print(" Type Y to quit. Type N to continue on the same segment. Type S to go back to the sentence level.")
            stopper = input()
            if stopper == "y":
                break
            if stopper == "n":
                helper_to_stopper = 0
                for i in range(100):
                    helper_to_stopper += 1
                    current_word= select_word_in_sentence + helper_to_stopper
                    print(split_sen[current_word])
                    print("Press A to quit this section else press anything to add the value to the current grouping")
                    new_value=input()
                    if new_value != "a":
                        zero_list[current_word]=starting_point
                    if new_value == "a":
                        break
            if stopper == "s":
                continue     

    print(str(zero_list))
    
    sheet[f'D{starting_row_value}'] = str(zero_list)
        

        
    
    wb.save(r'D:\Kimlichcova\sparrow_excelled.xlsx')
    
    