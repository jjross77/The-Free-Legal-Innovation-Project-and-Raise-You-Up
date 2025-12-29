# -*- coding: utf-8 -*-
"""
Created on Wed Jan 18 06:19:26 2023

@author: rossg
"""
import openpyxl
import json
#import ast could be a solution we might try another day

# don't need to reload column 4 to data because you can apply the tokenizer from nltk again on intital column and should be fine
# do this when feeding the input tokens with the labels for later processing. 

# Give the location of the file
path = r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Downloads\Sparrow_excell3.xlsx"
 
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
 
es = wb_obj.active
max_col = es.max_column
max_row= es.max_row 
#print(max_row)
#print(max_col)
 

column4_dic = {}
try_list = []
# Loop will print all columns name
# THIS ONE WORKS FOR THE NUMBER COLUMN LIST
#if cell_obj.value != None:
    #
    # USE THIS WHEN LOADING THE NUBMER ROW COLUMN 4 
    #x= json.loads(cell_obj.value)
    #column4_dic[i] = x
for i in range(2, 100):
    cell_obj = es.cell(row = i, column = 3)
    cell_obj2 = es.cell(row=i, column = 1 )
    
    print(cell_obj.value)
    print(cell_obj2.value)
    
    

