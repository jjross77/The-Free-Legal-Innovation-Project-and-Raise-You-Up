# -*- coding: utf-8 -*-
"""
Created on Mon Feb 20 01:30:48 2023

@author: rossg
"""

import multiprocessing

def run_tokenizer(start, stop):
    import torch
    from transformers import DebertaTokenizerFast
    import openpyxl
    from transformers import DebertaModel
    import re
    import pickle
    from openpyxl import load_workbook
    
    torch.set_printoptions(profile="full")
    tokenizer = DebertaTokenizerFast.from_pretrained("microsoft/deberta-base", truncation_side="left")
    #.half
    find_word_pattern = re.compile(r"\w+")
    finding_start_of_second_sentence=re.compile(r"1")
    finding_start_of_first_sentence= re.compile(r"0")

    path = r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Downloads\Sparrow_excell3.xlsx"
    


    column4_dic = {}
    try_list = []
    
    wb = load_workbook(filename=path)
    sheet = wb.active
    starting_value_minus2 = ""
    storing = 0

        
    for i, value in enumerate(sheet['B']):
        storing += 1
        if value.value == None:
            starting_value_minus2=value
            break 



    print(storing)    
    final_row_value= storing -1
    sentence_list = []
    print("starting to create sentence list")
    
    for i in range(start, stop+2):# I can use this to ensure that
    #only one case should be fed at a time
    #when generating the embeddings for the model to ensure embeddings are good
        cell_obj2 = sheet.cell(row=i, column = 1 )
        sentence = cell_obj2.value
        sentence.encode("utf-8") # this is convert the string to utf-8 encoding
        sentence_list.append(sentence)
        
        
    token_list = []
    amount_of_token_list = []
    existing_sentences_list = []
    print("starting to tokenize")


    for i, sentence_2 in enumerate(sentence_list):
        if i==0:
            existing_sentences= f"{sentence_2}"
        
        #print(existing_sentences)
        finding_words_to_remove_from_sentence_list = find_word_pattern.findall(existing_sentences)
        finding_words_to_remove_len =len(finding_words_to_remove_from_sentence_list)
        #print(finding_words_to_remove_len)
        
        if finding_words_to_remove_len >= 550:
            
            word_to_be_removed_for_next_iteration= finding_words_to_remove_from_sentence_list[50]
            
            print(word_to_be_removed_for_next_iteration)
            
            pattern_to_find_word_to_be_kept = re.compile(f"{word_to_be_removed_for_next_iteration}")
            
            identifier_for_removing_from_string=re.search(pattern_to_find_word_to_be_kept, string=existing_sentences)
            
            end_of_word_needed_segment_sentence= identifier_for_removing_from_string.span()[1]
            
            existing_sentences=existing_sentences[end_of_word_needed_segment_sentence:]
        if i==0:
            inputs = tokenizer(sentence_2, return_tensors="pt")
            strwoo=(str((inputs['token_type_ids'])))
            tokens_in_current_sentence= re.findall(finding_start_of_first_sentence, string=strwoo)
            amount_of_tokens=len(tokens_in_current_sentence)-2 # need to check how many tokens I need to remove here
            # I think that two should work here but still need to check
            
            amount_of_tokens = 0 # will need to write in code for if I =0 relating to this
            
            
            
            
        if i!=0:
            #print(existing_sentences)
            #print(sentence_2)
            inputs = tokenizer(existing_sentences,sentence_2, return_tensors="pt") 
            print(inputs['token_type_ids'])
            strwoo=(str((inputs['token_type_ids'])))
            tokens_in_current_sentence= re.findall(finding_start_of_second_sentence, string=strwoo)
            amount_of_tokens=len(tokens_in_current_sentence)  # for final token that will be labeled as 
            #a 1 still need to remove this or find a way to remove the final token
            # I think I did manage to do this in the end
            
            
            existing_sentences= existing_sentences + " " + f"{sentence_2}"
        token_list.append(inputs)
        amount_of_token_list.append(amount_of_tokens)
        existing_sentences_list.append(existing_sentences)
        
    """with open(r"D:\Kimlichcova\amount_of_token_list.pickle","wb") as f:
        pickle.dump(amount_of_token_list, f, pickle.HIGHEST_PROTOCOL )  
    with open(r"D:\Kimlichcova\sentence_list.pickle","wb") as f:
        pickle.dump(sentence_list, f, pickle.HIGHEST_PROTOCOL )
    with open(r"D:\Kimlichcova\token_list.pickle","wb") as f:
        pickle.dump(token_list, f, pickle.HIGHEST_PROTOCOL )
    with open(r"D:\Kimlichcova\existing_sentences_list.pickle","wb") as f:
        pickle.dump(existing_sentences_list, f, pickle.HIGHEST_PROTOCOL )"""
    return token_list, amount_of_token_list, sentence_list
    

def Run_Model(sentence_list=[], token_list=[], amount_of_token_list=[]):
    
    from transformers import DebertaModel
    import torch
    import subprocess
    import time
    import sys
    import re
    import pickle
    import multiprocessing
    list_of_tokens_2 = []
    torch.cuda.empty_cache()
    #print(torch.cuda.memory_summary(device=None, abbreviated=False))
    torch.set_printoptions(profile="full")
    torch.cuda.empty_cache()
    #device = torch.device("cuda")
    dictt_of_sentence_embeddings= {}
    
    
    #with open(r'D:\Kimlichcova\amount_of_token_list.pickle', 'rb') as f:
        #amount_of_token_list = pickle.load(f)
    #with open(r'D:\Kimlichcova\token_list.pickle', 'rb') as f:
        #token_list = pickle.load(f)
        #print(token_list)
    
        
            

    #with open(r'D:\Kimlichcova\sentence_list.pickle', 'rb') as f:
         #sentence_list = pickle.load(f)
         #data = [sentence_list.cuda() for sentence in sentence_list] 

    model = DebertaModel.from_pretrained("microsoft/deberta-base").to("cuda").half()

    for sentence_3, token, num_token in zip(sentence_list, token_list, amount_of_token_list):
            
            current_sentence = sentence_3
            cuda_token=token.to("cuda")
            print(token)
            output = model(**cuda_token)
            torch.cuda.empty_cache()
            embedding_with_extra_end_token=output[0][0][-num_token:]
            good_embeddings= embedding_with_extra_end_token[:-1]
            
            sentence_embedding= torch.mean(good_embeddings, dim=0)# need to test this seems fine

            #print(sentence_3.encode("utf-8"))
            dictt_of_sentence_embeddings[sentence_3]= sentence_embedding
            


            
            
   

        

    
        #result=subprocess.run([sys.executable, r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Downloads\h.py"], capture_output=True, text=True)
        #print("stdout:", result.stdout)
        #print("stderr:", result.stderr)


token_list, amount_of_token_list, sentence_list = run_tokenizer(2, 200)   
      
if __name__ == '__main__':
    p = multiprocessing.Process(target=Run_Model, args=(sentence_list, token_list, amount_of_token_list))
    p.start()
    #When we comment out the join method, the main process finishes before the child process.

    p.join()       
    print("hello world")       
    

    


#last_hidden_states = outputs.last_hidden_state
