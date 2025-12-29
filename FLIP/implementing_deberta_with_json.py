# -*- coding: utf-8 -*-
"""
Created on Thu Mar  2 23:02:56 2023

@author: yyyyyyyyyyyyyyyyyyyy
"""


import multiprocessing

def Run_model():
    
    set("PYTORCH_CUDA_ALLOC_CONF=garbage_collection_threshold:0.6,max_split_size_mb:24") 

    import torch
    from transformers import DistilBertTokenizerFast
    from transformers import DistilBertModel
    
    import re
    import pickle
    from openpyxl import load_workbook
    import os

    

    torch.set_printoptions(profile="full")
    

    #.half
    find_word_pattern = re.compile(r"\w+")
    finding_start_of_second_sentence=re.compile(r"1")
    finding_start_of_first_sentence= re.compile(r"0")

    path = r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Downloads\Sparrow_excell3.xlsx"
    
    

    sentence_list = []
    past_case_being_iterated_through =""
    case_start_list = [] 
    label_2_for_sentence_type_list = []
    label_1_for_good_or_not_list = []
    sentence_list_new= []
    label_1_for_good_or_not_list_new=[]
    label_2_for_sentence_type_list_new = []
    case_start_list_new = []
    



    
    wb = load_workbook(filename=path)
    sheet = wb.active
    #starting_value_minus2 = ""
    storing = 0
    counterr= 0
    past_case_numm=0
    tag=0
    previous_tag=0
    dictt_of_sentence_embeddings= {}
    current_sentence_value_in_list = 0
    last_case_looked_at=0
    exception_executed = "false"
    existing_sentences=""
    iiii= 0
    recording_index_values_for_current_sentence = []
    recording_index_values_for_current_sentence_now = 0

        
    for i, value in enumerate(sheet['B']):
        storing += 1
        if value.value == None:
            #starting_value_minus2=value
            break 
    storing= storing+1# might need to add one here, missed one of the sentences at the end
    # I do need to add one here
 
    
    for ii in range(2, storing):
        
        
        if os.path.exists(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\sentence_list.pickle") == False:#1
            cell_obj1 = sheet.cell(row=ii, column = 1)
            sentence = cell_obj1.value
            #sentence=sentence.encode("utf-8")
            sentence_list.append(sentence)
            #cell 2 is zero cell 3 is 1 so it is two ahead, and that should be right
            
            
        if os.path.exists(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\label_1_for_good_or_not_list.pickle") == False:#2
            cell_obj2 = sheet.cell(row=ii, column = 2 )
            label_1_for_good_or_not = cell_obj2.value
            label_1_for_good_or_not_list.append(label_1_for_good_or_not)
        

            

        if os.path.exists(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\label_2_for_sentence_type_list.pickle") == False:#3
            cell_obj3 = sheet.cell(row=ii, column = 3 )
            label_2_for_sentence_type = cell_obj3.value
            label_2_for_sentence_type_list.append(label_2_for_sentence_type)



        if os.path.exists(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\case_start_list.pickle") == False:#4
            cell_obj4 = sheet.cell(row=ii, column = 4 )
            current_case_being_iterated_through = cell_obj4.value
            if current_case_being_iterated_through != past_case_being_iterated_through:
                counterr+= 1
            case_start_list.append(counterr)
            past_case_being_iterated_through = current_case_being_iterated_through
            
        
            
        
        if os.path.exists(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\sentence_list.pickle") == True:#1
            with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\sentence_list.pickle","rb") as f1:
                sentence_list=pickle.load(f1) 

        

            
        if os.path.exists(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\label_1_for_good_or_not_list.pickle") == True:#2
            with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\label_1_for_good_or_not_list.pickle","rb") as f2:
                label_1_for_good_or_not_list=pickle.load(f2) 


            
        

        if os.path.exists(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\label_2_for_sentence_type_list.pickle") == True:#3
            with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\label_2_for_sentence_type_list.pickle","rb") as f3:
                label_2_for_sentence_type_list=pickle.load(f3) 



        if os.path.exists(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\case_start_list.pickle") == True:#4
            with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\case_start_list.pickle","rb") as f4:
                case_start_list=pickle.load(f4) 
                
        if os.path.exists(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\recording_index_values_for_current_sentence.pickle") == True:#5
            with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\recording_index_values_for_current_sentence.pickle","rb") as f5:
                recording_index_values_for_current_sentence=pickle.load(f5) 
                break
        
    
    
    
    if os.path.exists(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\case_start_list.pickle") == False:
        
        with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\sentence_list.pickle","wb") as f6: #1
            pickle.dump(sentence_list, f6, pickle.HIGHEST_PROTOCOL)
        with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\label_1_for_good_or_not_list.pickle","wb") as f7:#2
            pickle.dump(label_1_for_good_or_not_list, f7, pickle.HIGHEST_PROTOCOL)
        with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\label_2_for_sentence_type_list.pickle","wb") as f8:#3
            pickle.dump(label_2_for_sentence_type_list, f8, pickle.HIGHEST_PROTOCOL)
        with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\case_start_list.pickle","wb") as f9:#4
            pickle.dump(case_start_list, f9, pickle.HIGHEST_PROTOCOL)
        

    


    if os.path.exists(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\existing_senteces_and_current_sentence.pickle") == False:# if_ not existing_sentences
        sentence_list_new= sentence_list
        label_1_for_good_or_not_list_new = label_1_for_good_or_not_list
        label_2_for_sentence_type_list_new = label_2_for_sentence_type_list
        case_start_list_new = case_start_list
        




    if os.path.exists(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\existing_senteces_and_current_sentence_new.pickle") == True:# if_existing _sentences
            with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\existing_senteces_and_current_sentence_new.pickle","rb") as f9:
                existing_senteces_and_current_sentence=pickle.load(f9) 
            existing_sentences= existing_senteces_and_current_sentence[0]
            current_sentence= existing_senteces_and_current_sentence[1]
            #since we added the sentence accdientialy to existing_sentences we need to remove it so do below
            length_of_current_sentence=len(current_sentence)
            existing_sentences=existing_sentences[:-length_of_current_sentence]
            #this is the problem
            # we need to make it find the right current sentnece
            #it needs to match like the existing_sentences
            #need to check that the sentence or sentences before it match existing
            
            
            #what if there is start of new case when it fails? need to cosndier this
            #for the code below
            current_sentence_value_in_list= sum(recording_index_values_for_current_sentence)
            
            
           
            

            sentence_list_new= sentence_list[current_sentence_value_in_list:]
            label_1_for_good_or_not_list_new=label_1_for_good_or_not_list[current_sentence_value_in_list:]
            label_2_for_sentence_type_list_new = label_2_for_sentence_type_list[current_sentence_value_in_list:]
            
            case_start_list_new = case_start_list[current_sentence_value_in_list:]
            
            last_case_looked_at=current_sentence_value_in_list-1
            
            past_case_numm=case_start_list[last_case_looked_at]
            #print(existing_sentences)
            #print(past_case_numm)
            #print(type(past_case_numm))
            #print(existing_sentences)
            print(sentence_list_new[0])
            """
    
            del sentence_list
            del label_1_for_good_or_not_list
            del label_2_for_sentence_type_list
            del case_start_list
            del wb 
            del sheet
            del storing
            del counterr
            """
            

    
    
    
    
    tokenizer = DistilBertTokenizerFast.from_pretrained("distilbert-base-uncased", truncation_side="left")
    model = DistilBertModel.from_pretrained("distilbert-base-uncased").to("cuda").half()
    #print(len(sentence_list_new))
    #print(len(case_start_list_new))
    #print(len(label_2_for_sentence_type_list_new))
    #print(len(label_1_for_good_or_not_list_new))
    
    
    
    
    for sentence_2, case_numm, label2, label1 in zip(sentence_list_new, case_start_list_new, label_2_for_sentence_type_list_new, label_1_for_good_or_not_list_new):
        current_case_numm = case_numm
        recording_index_values_for_current_sentence_now+=1 # this is how we will find the start of the case that we are looking for
        
        #print(sentence_2)
        
        if current_case_numm!= past_case_numm:
            existing_sentences= f"{sentence_2}"
            
            
        finding_words_to_remove_from_sentence_list = find_word_pattern.findall(existing_sentences)
        finding_words_to_remove_len =len(finding_words_to_remove_from_sentence_list)
        #print(finding_words_to_remove_len)
        
        if finding_words_to_remove_len > 550:
            
            
            word_to_be_removed_for_next_iteration= finding_words_to_remove_from_sentence_list[-550]
            
            #print(word_to_be_removed_for_next_iteration)
            #print(word_to_be_removed_for_next_iteration)
            
            pattern_to_find_word_to_be_kept = re.compile(f"{word_to_be_removed_for_next_iteration}")
            
            identifier_for_removing_from_string=re.search(pattern_to_find_word_to_be_kept, string=existing_sentences)
            
            end_of_word_needed_segment_sentence= identifier_for_removing_from_string.span()[1]
            existing_sentences=existing_sentences[end_of_word_needed_segment_sentence:]
            
            
        if current_case_numm != past_case_numm:
            tag+=1
            inputs = tokenizer(sentence_2, return_tensors="pt", truncation=True, return_token_type_ids=True)
            #print(inputs)
            strwoo=(str((inputs['token_type_ids'])))
            tokens_in_current_sentence= re.findall(finding_start_of_first_sentence, string=strwoo)
            amount_of_tokens=len(tokens_in_current_sentence)-2 # need to check how many tokens I need to remove here
            # I think that two should work here but still need to check
            #print(amount_of_tokens)
            inputs = tokenizer(sentence_2, return_tensors="pt", truncation=True).to("cuda")
            #amount_of_tokens = 0 # will need to write in code for if I =0 relating to this
            
            
            
        
        if  current_case_numm == past_case_numm:
            
            
            #print(existing_sentences)
            #print(sentence_2)
            inputs = tokenizer(existing_sentences, sentence_2, return_tensors="pt", truncation=True, return_token_type_ids=True)
            #print(inputs['token_type_ids'])
            strwoo=(str((inputs['token_type_ids'])))
            tokens_in_current_sentence= re.findall(finding_start_of_second_sentence, string=strwoo)
            amount_of_tokens=len(tokens_in_current_sentence)  # for final token that will be labeled as 
            #print(amount_of_tokens)
            #a 1 still need to remove this or find a way to remove the final token
            # I think I did manage to do this in the end
            inputs = tokenizer(existing_sentences, sentence_2, return_tensors="pt", truncation=True).to("cuda")    
            
            
            
            existing_sentences= existing_sentences + " " + f"{sentence_2}"
        
        
            
            
                
     
        
        #existing_sentences.encode("utf-8")
        past_case_numm = current_case_numm
        
        
        
        
        with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\existing_sentence_list_5.txt","a" , encoding="utf-8" ) as fff:
            fff.write(f"\n {existing_sentences} \n")
           
            
        try:

            output = model(**inputs)
            
            if tag > previous_tag:
                embedding_with_extra_start_and_end_token=output[0][0]
                good_embeddings= embedding_with_extra_start_and_end_token[1:]
                good_embeddings= good_embeddings[:-1]
                sentence_embedding= torch.mean(good_embeddings, dim=0)# need to test this seems fine
            
            else:
                embedding_with_extra_end_token=output[0][0][-amount_of_tokens:]
                good_embeddings= embedding_with_extra_end_token[:-1]
                
                sentence_embedding= torch.mean(good_embeddings, dim=0)# need to test this seems fine
                
                #print(sentence_2.encode("utf-8"))
            sentence_3= sentence_2
            if len(sentence_2)>600:
                sentence_3 = sentence_2[:600]
            # this is to allow us to retain sentence 2 and also let us truncate the dict
            
            iiii+= 1
            # need to find a way to make sure every embedding is unique
            # with this technique when combinding dictionaries could lose data, very slim but possible
            # maybe need to do another thing when combining the dictioanries to ensure all keys unique
            # add additonal value to key name
            
            dictt_of_sentence_embeddings[f"{label1}: {label2}: {iiii}: {sentence_3}"]= sentence_embedding
            previous_tag = tag


            #cuda_token=token.to("cuda")
            torch.cuda.empty_cache()
        
        
        
        except:# Exception as e:
            print(existing_sentences)
            print(sentence_2)
            #print(recording_index_values_for_current_sentence)
            
            recording_index_values_for_current_sentence.append(recording_index_values_for_current_sentence_now-1)
            # by minusing one here this should fix the fact that we start at zero when we iteratre through
            
            with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\recording_index_values_for_current_sentence.pickle","wb") as f10:
                pickle.dump(recording_index_values_for_current_sentence, f10, pickle.HIGHEST_PROTOCOL)
            
            existing_senteces_and_current_sentence_new = [existing_sentences, sentence_2]
    
    
            #print(current_sentence) # need to learn how to use Json instead to store the dicitonary
            with open(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\existing_senteces_and_current_sentence_new.pickle","wb") as f11:
                pickle.dump(existing_senteces_and_current_sentence_new, f11, pickle.HIGHEST_PROTOCOL)
            #save current sentence and existing sentence
            dir_of_batches=os.listdir(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\embedding_batches")
            doc_number=len(dir_of_batches)
            
            embedding_file_name= f"dicttt_of_sentence_embeddings{doc_number}.pickle"
            embedding_path_name= r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\embedding_batches" + "\\" + embedding_file_name
            with open(embedding_path_name,"wb") as f11:
                pickle.dump(dictt_of_sentence_embeddings, f11, pickle.HIGHEST_PROTOCOL)
   
            #print(e)
            exception_executed= "true"
            break
    if exception_executed!="true":
        dir_of_batches=os.listdir(r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\embedding_batches")
        doc_number=len(dir_of_batches)
        
        embedding_file_name= f"dicttt_of_sentence_embeddings{doc_number}.pickle"
        embedding_path_name= r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Documents\embedding_batches" + "\\" + embedding_file_name
        with open(embedding_path_name,"wb") as f12:
            pickle.dump(dictt_of_sentence_embeddings, f12, pickle.HIGHEST_PROTOCOL)
    
        
    
 
   
#Run_model()
 

for i in range(270):
    if __name__ == '__main__':
        p = multiprocessing.Process(target=Run_model)
        p.start()
        #When we comment out the join method, the main process finishes before the child process.

        p.join()       
        print("hello world")   