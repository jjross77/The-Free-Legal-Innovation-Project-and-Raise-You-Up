# -*- coding: utf-8 -*-
"""
Created on Wed Jan 25 02:52:21 2023

@author: rossg
"""

import numpy as np

from nltk.tokenize import word_tokenize


#Useful v Non-Useful

real_sentence = 1
non_sentence = 2



# FACTS

actual_facts = 1
lower_court_arg_parties = 2
lower_court_non_arg = 3
facts = actual_facts + lower_court_arg_parties + lower_court_non_arg

# Analysis

interpretation_analysis = 4
quotes_analysis = 5
tests_analysis = 6

analysis = interpretation_analysis + quotes_analysis + tests_analysis

#Issues

issue = 7
lower_court_issue = 8
issues = issue + lower_court_issue 


# Holdings
lower_court_holding = 9
current_court_holding = 10
holdings = lower_court_holding + current_court_holding

#token values

legislation = 1
jurisprudence = 2
articles = 3
coresponding_page_num = 4
references_leg = 5
references_juris = 6
references_article= 7

#Token Values for Interpretation Sentences
main_subject_being_interpreted = 1


# Type of Legal citiation

concurring = 1
varied_differentiated = 2 
overruled = 3


import openpyxl

from openpyxl import load_workbook
wb = load_workbook(filename=r'E:\Kimlichcova\Sparrow_excell3.xlsx')
sheet = wb.active
Column_A = sheet['A'] 
Column_B = sheet['B']  # length that we start labeling again when we reopen the file. 
leng_of_columna= len(Column_A)
leng_of_columnB= len(Column_B)
wb = load_workbook(filename=r'E:\Kimlichcova\Sparrow_excell3.xlsx')
sheet = wb.active
starting_value_minus2 = ""
storing = 0
for i, value in enumerate(sheet['B']):
    storing += 1
    if value.value == None:
        starting_value_minus2=value
        break  
starting_row_value= storing -1
print(starting_row_value)


helper_to_stopper = 0   
temp_list = []
sen_wait = 0
for i, sentence in enumerate(range(leng_of_columna)):
    temp_list.clear()
    starting_row_value += 1
    c = sheet[f'A{starting_row_value}']
    split_sen=word_tokenize(c.value)
    print(c.value)
    len_sentence=len(split_sen)
    zero_list= [0 for i in range(len_sentence)]
    zero_list2= [0 for i in range(len_sentence)]

    for i, word in enumerate(split_sen):
        temp_item =f"{word}_({i})"
        temp_list.append(temp_item)

    label1= 1
    sheet[f'B{starting_row_value}'] = label1
    
    if i % 2 == 0:
        print("""Useful Facts = 1,
          useful_analysis = 2,
          all_holdings =3, 
          all_Issues = 4, all_rules = 4,
          conclusion_like_analysis_by_court = 5, 
          lower_court arguments, and eveyrone else analysis =6,
          descriptive_sentences = 7, N/A = 8, 
          Category for broken_bits = 9,  10, Quotes = 10""")

    label2= int(input())
    sheet[f'C{starting_row_value}'] = label2
    if label2==8:
        sheet[f'B{starting_row_value}'] = 2
        
    
    
    
    
        
        
        
    
    wb.save(r'E:\Kimlichcova\Sparrow_excell3.xlsx')
    
    