# -*- coding: utf-8 -*-
"""
Created on Wed Feb 22 17:30:16 2023

@author: yyyyyyyyyyyyyyyyyyyy
"""


import multiprocessing

def Run_model():
    import torch
    from transformers import DebertaTokenizerFast
    import openpyxl
    from transformers import DebertaModel
    import re
    import pickle
    from openpyxl import load_workbook
    import os

    

    torch.set_printoptions(profile="full")
    tokenizer = DebertaTokenizerFast.from_pretrained("microsoft/deberta-base", truncation_side="left")
    model = DebertaModel.from_pretrained("microsoft/deberta-base").to("cuda").half()

    #.half
    find_word_pattern = re.compile(r"\w+")
    finding_start_of_second_sentence=re.compile(r"1")
    finding_start_of_first_sentence= re.compile(r"0")

    path = r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Downloads\Sparrow_excell3.xlsx"
    
    


    column4_dic = {}
    try_list = []
    past_case_being_iterated_through =""
    case_start_list = []
    sentence_dict = {}
    iterable_list = []   



    
    wb = load_workbook(filename=path)
    sheet = wb.active
    starting_value_minus2 = ""
    storing = 0
    counterr= 0
    past_case_numm=""
    num_of_embeddings= 0
    

        
    for i, value in enumerate(sheet['B']):
        storing += 1
        if value.value == None:
            starting_value_minus2=value
            break 
        



    print(storing)    
    final_row_value= storing -1
    sentence_list = []
    print("starting to create sentence list")
    #marking where one case starts and another one stops
    
    if os.path.exists(r"D:\Kimlichcova\dictt_of_sentence_embeddings.pickle") == True:
        with open(r"D:\Kimlichcova\dictt_of_sentence_embeddings.pickle","rb") as ff:
            dict_of_sentence_embeddings_loaded=pickle.load(ff) 
            num_of_embeddings=len(dict_of_sentence_embeddings_loaded) -1
            # the reason we are adding one here is to maybe solve the issue
            #when in the big loop there will be a point
            # where zero and the value need <= and this may solve it
            #if we add 1 this means that the iteration after the embedding will be used
            # this is our starting value
            
        
    """if os.path.exists(r"D:\Kimlichcova\dictt_of_sentence_embeddings.pickle") != True:
        start"""
        
        
    for i in range(2, final_row_value):
        
        # I can use this to ensure that
    #only one case should be fed at a time
    #when generating the embeddings for the model to ensure embeddings are good
        cell_obj2 = sheet.cell(row=i, column = 4 )
        cell_obj1 = sheet.cell(row=i, column = 1 )
        sentence = cell_obj1.value
        sentence.encode("utf-8") # this is convert the string to utf-8 encoding
        current_case_being_iterated_through = cell_obj2.value
        #current_case_being_iterated_through.encode("utf-8") # this is convert the string to utf-8 encoding
        if current_case_being_iterated_through != past_case_being_iterated_through:
            counterr+= 1
            
            #mark value in sentence_list somehow
            
        past_case_being_iterated_through = current_case_being_iterated_through
        sentence_list.append(sentence)
        case_start_list.append(counterr)
        

        
        
        
        
        

    """for i in range(2, stop+2):
        # I can use this to ensure that
    #only one case should be fed at a time
    #when generating the embeddings for the model to ensure embeddings are good
        cell_obj2 = sheet.cell(row=i, column = 1 )
        sentence = cell_obj2.value
        sentence.encode("utf-8") # this is convert the string to utf-8 encoding
        sentence_list.append(sentence)"""
        
        
    #token_list = []
    #amount_of_token_list = []
    #existing_sentences_list = []
    dictt_of_sentence_embeddings= {}



    print("starting to tokenize")
    for i, sentence_2, case_numm in zip(range(len(sentence_list)),sentence_list, case_start_list):
        print(i)
        current_case_numm = case_numm
           #num_of_embedding becomes number of iterations of loop skipped before continuing but 
            #still need existing_sentences to be generated
        
        if i==0 or current_case_numm!= past_case_numm:
            existing_sentences= f"{sentence_2}"
        
        #print(existing_sentences)
        finding_words_to_remove_from_sentence_list = find_word_pattern.findall(existing_sentences)
        finding_words_to_remove_len =len(finding_words_to_remove_from_sentence_list)
        #print(finding_words_to_remove_len)
        
        if finding_words_to_remove_len >= 550:
            
            word_to_be_removed_for_next_iteration= finding_words_to_remove_from_sentence_list[50]
            
            print(word_to_be_removed_for_next_iteration)
            
            pattern_to_find_word_to_be_kept = re.compile(f"{word_to_be_removed_for_next_iteration}")
            
            identifier_for_removing_from_string=re.search(pattern_to_find_word_to_be_kept, string=existing_sentences)
            
            end_of_word_needed_segment_sentence= identifier_for_removing_from_string.span()[1]
            
            existing_sentences=existing_sentences[end_of_word_needed_segment_sentence:]
            
        if i==0 or current_case_numm != past_case_numm:
            if i > num_of_embeddings:
                inputs = tokenizer(sentence_2, return_tensors="pt", truncation=True).to("cuda")
                strwoo=(str((inputs['token_type_ids'])))
                tokens_in_current_sentence= re.findall(finding_start_of_first_sentence, string=strwoo)
                amount_of_tokens=len(tokens_in_current_sentence)-2 # need to check how many tokens I need to remove here
                # I think that two should work here but still need to check
                
                #amount_of_tokens = 0 # will need to write in code for if I =0 relating to this

                
            
            
            
            
            
            
        if i!=0 or current_case_numm == past_case_numm:
            #print(existing_sentences)
            #print(sentence_2)
            if i > num_of_embeddings:
                inputs = tokenizer(existing_sentences,sentence_2, return_tensors="pt", truncation=True).to("cuda") 
                #print(inputs['token_type_ids'])
                strwoo=(str((inputs['token_type_ids'])))
                tokens_in_current_sentence= re.findall(finding_start_of_second_sentence, string=strwoo)
                amount_of_tokens=len(tokens_in_current_sentence)  # for final token that will be labeled as 
                #a 1 still need to remove this or find a way to remove the final token
                # I think I did manage to do this in the end
            
            
            
            existing_sentences= existing_sentences + " " + f"{sentence_2}"
        #token_list.append(inputs)
        #amount_of_token_list.append(amount_of_tokens)
        #existing_sentences_list.append(existing_sentences)
        #current_sentence = sentence_2
        past_case_numm = current_case_numm
        if i > num_of_embeddings:
            
            try:
                
                output = model(**inputs)
                embedding_with_extra_end_token=output[0][0][-amount_of_tokens:]
                good_embeddings= embedding_with_extra_end_token[:-1]
                
                sentence_embedding= torch.mean(good_embeddings, dim=0)# need to test this seems fine

                print(sentence_2.encode("utf-8"))
                
                dictt_of_sentence_embeddings[sentence_2]= sentence_embedding
                

                
                #cuda_token=token.to("cuda")
                torch.cuda.empty_cache()
            
            except:
                
                #print(current_sentence) # need to learn how to use Json instead to store the dicitonary
                if os.path.exists(r"D:\Kimlichcova\dictt_of_sentence_embeddings.pickle") == True:
                    with open(r"D:\Kimlichcova\dictt_of_sentence_embeddings.pickle","rb") as f:
                        original_case_dict=pickle.load(f)
                        for keys, values in original_case_dict.items():
                            dictt_of_sentence_embeddings[keys]= values
                    with open(r"D:\Kimlichcova\dictt_of_sentence_embeddings.pickle","wb") as f:
                        pickle.dump(dictt_of_sentence_embeddings, f, pickle.HIGHEST_PROTOCOL) 
                if os.path.exists(r"D:\Kimlichcova\dictt_of_sentence_embeddings.pickle") != True:
                    with open(r"D:\Kimlichcova\dictt_of_sentence_embeddings.pickle","wb") as f:
                        pickle.dump(dictt_of_sentence_embeddings, f, pickle.HIGHEST_PROTOCOL) 
                break 
            
       
        
                

            
            
           

            
            

        
        
        
        
    """with open(r"D:\Kimlichcova\amount_of_token_list.pickle","wb") as f:
        pickle.dump(amount_of_token_list, f, pickle.HIGHEST_PROTOCOL )  
    with open(r"D:\Kimlichcova\sentence_list.pickle","wb") as f:
        pickle.dump(sentence_list, f, pickle.HIGHEST_PROTOCOL )
    with open(r"D:\Kimlichcova\token_list.pickle","wb") as f:
        pickle.dump(token_list, f, pickle.HIGHEST_PROTOCOL )
    with open(r"D:\Kimlichcova\existing_sentences_list.pickle","wb") as f:
        pickle.dump(existing_sentences_list, f, pickle.HIGHEST_PROTOCOL )"""
    #list_of_tokens_2 = []
    #torch.cuda.empty_cache()
    #print(torch.cuda.memory_summary(device=None, abbreviated=False))
    #torch.cuda.empty_cache()
    #device = torch.device("cuda")
    
    
    #with open(r'D:\Kimlichcova\amount_of_token_list.pickle', 'rb') as f:
        #amount_of_token_list = pickle.load(f)
    #with open(r'D:\Kimlichcova\token_list.pickle', 'rb') as f:
        #token_list = pickle.load(f)
        #print(token_list)
    
        
            

    #with open(r'D:\Kimlichcova\sentence_list.pickle', 'rb') as f:
         #sentence_list = pickle.load(f)
         #data = [sentence_list.cuda() for sentence in sentence_list] 


            
            
                

    
   


#Run_model()   

for i in range(500):
    if __name__ == '__main__':
        p = multiprocessing.Process(target=Run_model)
        p.start()
        #When we comment out the join method, the main process finishes before the child process.

        p.join()       
        print("hello world")     
  
   

    

#p = multiprocessing.Process(target=Run_model, args=(2,200))
#last_hidden_states = outputs.last_hidden_state
