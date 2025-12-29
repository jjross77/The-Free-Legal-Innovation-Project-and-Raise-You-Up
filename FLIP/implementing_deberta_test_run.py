# -*- coding: utf-8 -*-
"""
Created on Wed Feb 15 04:10:02 2023

@author: yyyyyyyyyyyyyyyyyyyy
"""
import multiprocessing

def run_model():
    from transformers import  DebertaModel
    import torch
    from transformers import DebertaTokenizerFast

    import openpyxl
    import json
    import subprocess
    import time
    import sys
    torch.cuda.empty_cache()

    device = torch.device("cuda")
    tokenizer = DebertaTokenizerFast.from_pretrained("microsoft/deberta-base", truncation_side="left")
    model = DebertaModel.from_pretrained("microsoft/deberta-base").to("cuda")
    #.half
    print(torch.cuda.memory_summary(device=None, abbreviated=False))


    #import ast could be a solution we might try another day

    # don't need to reload column 4 to data because you can apply the tokenizer from nltk again on intital column and should be fine
    # do this when feeding the input tokens with the labels for later processing. 

    # Give the location of the file
    path = r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Downloads\Sparrow_excell3.xlsx"
     
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
     
    es = wb_obj.active
    max_col = es.max_column
    max_row= es.max_row 



    column4_dic = {}
    try_list = []
    # Loop will print all columns name
    # THIS ONE WORKS FOR THE NUMBER COLUMN LIST
    #if cell_obj.value != None:
        #
        # USE THIS WHEN LOADING THE NUBMER ROW COLUMN 4 
        #x= json.loads(cell_obj.value)
        #column4_dic[i] = x
        
    import openpyxl

    from openpyxl import load_workbook
    wb = load_workbook(filename=path)
    sheet = wb.active
    Column_A = sheet['A'] 
    Column_B = sheet['B']  # length that we start labeling again when we reopen the file. 
    leng_of_columna= len(Column_A)
    leng_of_columnB= len(Column_B)
    wb = load_workbook(filename=path)
    sheet = wb.active
    starting_value_minus2 = ""
    storing = 0

        
    for i, value in enumerate(sheet['B']):
        storing += 1
        if value.value == None:
            starting_value_minus2=value
            break 



    print(storing)    
    final_row_value= storing -1
    dictt_of_sentence = {}
    sentence_list = []
    for i in range(30, final_row_value+2):
        cell_obj2 = sheet.cell(row=i, column = 1 )
        sentence = cell_obj2.value
        sentence_list.append(sentence)
    token_list = []

    for i in enumerate(sentence_list):
        inputs = tokenizer(sentence, return_tensors="pt").to("cuda")   
        token_list.append(inputs)

    try:
        for i, sentence in enumerate(sentence_list):
            current_sentence= sentence
            output = model(**inputs)
            dictt_of_sentence[sentence]= output
            torch.cuda.empty_cache()
            print(f"{i}")
            #torch.mean()

            
            
    except:
        where_we_left_off= current_sentence
        print(where_we_left_off)
        multiprocessing.terminate() 

        #result=subprocess.run([sys.executable, r"C:\Users\yyyyyyyyyyyyyyyyyyyy\Downloads\h.py"], capture_output=True, text=True)
        #print("stdout:", result.stdout)
        #print("stderr:", result.stderr)

if __name__ == '__main__':
    p = multiprocessing.Process(target=run_model)
    p.start()
    #When we comment out the join method, the main process finishes before the child process.

    p.join()       
    print("hello world")
            
run_model()            
    

    
    
    
    




    
    




#last_hidden_states = outputs.last_hidden_state
